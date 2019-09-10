import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb['Sheet'] # or wb.active
n = int(input('enter n| '))

boldFont = Font(bold=True)

# for n = 3
#
#		1	2	3
#	1	1	2	3
#	2	2	4	6
#	3	3	6	9	

for i in range(2, 2+n):
	sheet['A' + str(i)].value = i-1
	sheet['A' + str(i)].font = boldFont
	sheet[get_column_letter(i) + '1'].value = i-1
	sheet[get_column_letter(i) + '1'].font = boldFont

for i in range(2, sheet.max_row + 1):
	for j in range(2, sheet.max_column+1):
		sheet[get_column_letter(j) + str(i)].value = sheet[get_column_letter(j) + '1'].value * sheet['A' + str(i)].value

wb.save('multTable.xlsx')