import openpyxl, os
from openpyxl.utils import get_column_letter, column_index_from_string

wkbk = str(input('enter workbook name| '))
wb = openpyxl.load_workbook(wkbk)
sheet = wb.active
file_name = str(input('enter filename| '))
file = open(file_name)
content = file.readlines()

line_count = len(content)
print(sheet.max_column)
if sheet.max_column == 1:
	print(sheet['A1'].value)
	if sheet['A1'].value == None:
		print('new col is 1 as A1 is empty')
		new_col = sheet.max_column
	else:
		print('new col is 2')
		new_col = sheet.max_column + 1
else:
	print('new col is new')
	new_col = sheet.max_column + 1

print(new_col)
for i in range(1, line_count+1):
	sheet[get_column_letter(new_col) + str(i)].value = content[i-1]
wb.save(wkbk)
