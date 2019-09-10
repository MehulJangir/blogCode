import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wkbk = str(input('enter workbook name| '))
wb = openpyxl.load_workbook(wkbk)
sheet = wb.active
# file_name = str(input('enter output file name| '))
# out_file = open(file_name, 'w')


if sheet['A1'].value == None:
	print('spreadsheet empty')
else:
	n_cols = sheet.max_column
	for col in range(1, n_cols+1):
		file_name = str(input('enter output file name for column {}| '.format(get_column_letter(col))))
		out_file = open(file_name, 'w')
		row = 1
		while sheet[get_column_letter(col) + str(row)].value != None:
			out_file.write(sheet[get_column_letter(col) + str(row)].value + '\n')
			print(sheet[get_column_letter(col) + str(row)].value)
			row = row + 1
		out_file.close()
