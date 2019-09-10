import sys, openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

if len(sys.argv) == 4:
	wb = openpyxl.load_workbook(sys.argv[3]) # or sys.argv[-1]
	sheet = wb.active
	start = int(sys.argv[1])
	skip = int(sys.argv[2])
	for i in range(sheet.max_row, start-1, -1):
		for j in range(1, sheet.max_column+1):		
			sheet[get_column_letter(j) + str(i+skip)].value = sheet[get_column_letter(j) + str(i)].value
	for i in range(start, start+skip):
		for j in range(1, sheet.max_column+1):
			sheet[get_column_letter(j)+str(i)].value = ''

	wb.save('test.xlsx')
else:
	print('insufficient number of inputs')