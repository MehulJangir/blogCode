import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wkbk = str(input('enter workbook name| '))
wb = openpyxl.load_workbook('avengers.xlsx')
wb_inv = openpyxl.Workbook()
sheet = wb.active
sheet_inv = wb_inv.active
data = list()

for i in range(1, sheet.max_row+1):
	data.append([])
	for j in range(1, sheet.max_column+1):
		data[i-1].append(sheet[get_column_letter(j) + str(i)].value)

for row in range(len(data)):
	for col in range(len(data[0])):
		sheet_inv[get_column_letter(row+1) + str(col+1)].value = data[row][col]

wb_inv.save('test.xlsx')